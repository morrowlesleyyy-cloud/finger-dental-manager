"""MEYA Dental Management System - Main Application"""
import os
import sys
import json
import uuid
from datetime import datetime, date, timedelta
from flask import (Flask, render_template, request, jsonify, redirect,
                   url_for, flash, Response, send_from_directory, session as flask_session,
                   make_response)
from sqlalchemy import func, extract, or_
from functools import wraps

# Add parent dir for imports
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

app = Flask(__name__)
app.secret_key = os.environ.get('SECRET_KEY', 'meya-dental-dev-key')

# Database - support DATABASE_PATH env for Docker
_env_db = os.environ.get('DATABASE_PATH')
if _env_db:
    db_path = _env_db
    os.makedirs(os.path.dirname(db_path), exist_ok=True)
else:
    db_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'clinic.db')
app.config['SQLALCHEMY_DATABASE_URI'] = f'sqlite:///{db_path}?timeout=30'
app.config['SQLALCHEMY_ENGINE_OPTIONS'] = {
    'connect_args': {'timeout': 30},
}
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False

from models import db, Patient, Appointment, Transaction, Employee
db.init_app(app)

# =================== Auth Configuration ===================
# Change these via environment variables
ADMIN_USERNAME = os.environ.get('ADMIN_USERNAME', 'admin')
ADMIN_PASSWORD = os.environ.get('ADMIN_PASSWORD', 'admin123')

# Store current user info in flask_session keys:
# logged_in, username, user_role, user_id
# role: 'admin' or 'staff'

def login_required(f):
    """Require valid session for route."""
    @wraps(f)
    def decorated(*args, **kwargs):
        if not flask_session.get('logged_in'):
            # API routes return 401, page routes redirect
            if request.path.startswith('/api/'):
                return jsonify({'error': 'unauthorized'}), 401
            return redirect(url_for('login_page'))
        return f(*args, **kwargs)
    return decorated

def normalize_phone(phone):
    """Ensure phone starts with '60' (Malaysia country code)."""
    if not phone:
        return None
    if not phone.startswith('60'):
        return '60' + phone
    return phone


def no_cache(response):
    """Add headers to prevent caching of protected pages."""
    if isinstance(response, str):
        response = make_response(response)
    response.headers['Cache-Control'] = 'no-store, no-cache, must-revalidate, max-age=0'
    response.headers['Pragma'] = 'no-cache'
    response.headers['Expires'] = '0'
    return response


# Global auth check - protect all routes except login and static
PUBLIC_PATHS = ['/login', '/api/login', '/api/logout', '/static/', '/api/auth/check', '/api/health', '/health']

@app.before_request
def check_global_auth():
    if request.method == 'OPTIONS':
        return None
    path = request.path
    # Allow public paths
    for public in PUBLIC_PATHS:
        if path == public or path.startswith(public):
            return None
    # Check session
    if not flask_session.get('logged_in'):
        if path.startswith('/api/'):
            return jsonify({'error': 'unauthorized'}), 401
        return redirect(url_for('login_page'))

# =================== SSE 客户端管理 ===================
sse_clients = []

def sse_broadcast(event, data):
    """推送 SSE 事件给所有连接的客户端"""
    msg = f'event: {event}\ndata: {json.dumps(data, ensure_ascii=False)}\n\n'
    dead = []
    for queue in sse_clients:
        try:
            queue.put(msg)
        except Exception:
            dead.append(queue)
    for q in dead:
        sse_clients.remove(q)

# =================== 创建表 (auto-import handles this) ===================

# =================== Auth Routes ===================

@app.route('/login', methods=['GET'])
def login_page():
    if flask_session.get('logged_in'):
        return redirect(url_for('index'))
    return render_template('login.html')


@app.route('/api/login', methods=['POST'])
def api_login():
    data = request.get_json()
    username = data.get('username', '')
    password = data.get('password', '')
    
    # Check admin (super user)
    if username == ADMIN_USERNAME and password == ADMIN_PASSWORD:
        flask_session['logged_in'] = True
        flask_session['username'] = username
        flask_session['user_role'] = 'admin'
        flask_session['user_id'] = 0
        flask_session['display_name'] = username
        flask_session['permissions'] = {
            'can_view_appointments': True,
            'can_view_transactions': True,
            'can_edit_patients': True,
            'is_admin': True,
        }
        return jsonify({'success': True, 'role': 'admin'})
    
    # Check employee accounts
    emp = Employee.query.filter_by(username=username, is_active=True).first()
    if emp and emp.password == password:
        flask_session['logged_in'] = True
        flask_session['username'] = emp.username
        flask_session['user_role'] = emp.role
        flask_session['user_id'] = emp.id
        flask_session['display_name'] = emp.display_name or emp.username
        flask_session['permissions'] = {
            'can_view_appointments': emp.can_view_appointments,
            'can_view_transactions': emp.can_view_transactions,
            'can_edit_patients': emp.can_edit_patients,
            'is_admin': emp.role == 'admin',
        }
        return jsonify({'success': True, 'role': emp.role})
    
    return jsonify({'success': False, 'error': '用户名或密码错误'}), 401


@app.route('/api/logout', methods=['POST'])
def api_logout():
    flask_session.clear()
    return jsonify({'success': True})


@app.route('/api/auth/check')
def api_auth_check():
    return jsonify({
        'logged_in': flask_session.get('logged_in', False),
        'username': flask_session.get('username', ''),
        'display_name': flask_session.get('display_name', ''),
        'role': flask_session.get('user_role', ''),
        'permissions': flask_session.get('permissions', {}),
    })


# =================== Employee Management API ===================

@app.route('/employees')
def employees_page():
    return no_cache(render_template('base.html'))


@app.route('/api/employees')
def api_list_employees():
    # Admin check
    if flask_session.get('user_role') != 'admin':
        return jsonify({'error': 'forbidden', 'message': '仅管理员可执行此操作'}), 403
    employees = Employee.query.order_by(Employee.id).all()
    return jsonify({'data': [e.to_dict() for e in employees]})


@app.route('/api/employees', methods=['POST'])
def api_create_employee():
    if flask_session.get('user_role') != 'admin':
        return jsonify({'error': 'forbidden', 'message': '仅管理员可执行此操作'}), 403
    data = request.get_json()
    username = data.get('username', '').strip()
    if not username:
        return jsonify({'error': '请输入用户名'}), 400
    password = data.get('password', '').strip()
    if not password:
        return jsonify({'error': '请输入密码'}), 400
    
    if Employee.query.filter_by(username=username).first():
        return jsonify({'error': '用户名已存在'}), 400
    
    emp = Employee(
        username=username,
        password=password,
        display_name=data.get('display_name', ''),
        role='staff',
        can_view_appointments=data.get('can_view_appointments', True),
        can_view_transactions=data.get('can_view_transactions', True),
        can_edit_patients=data.get('can_edit_patients', True),
    )
    db.session.add(emp)
    db.session.commit()
    return jsonify({'success': True, 'data': emp.to_dict()})


@app.route('/api/employees/<int:eid>', methods=['PUT'])
def api_update_employee(eid):
    if flask_session.get('user_role') != 'admin':
        return jsonify({'error': 'forbidden', 'message': '仅管理员可执行此操作'}), 403
    emp = Employee.query.get_or_404(eid)
    data = request.get_json()
    
    if 'username' in data:
        u = data['username'].strip()
        if u and u != emp.username:
            if Employee.query.filter_by(username=u).first():
                return jsonify({'error': '用户名已存在'}), 400
            emp.username = u
    if 'password' in data and data['password'].strip():
        emp.password = data['password'].strip()
    if 'display_name' in data:
        emp.display_name = data['display_name']
    for field in ['can_view_appointments', 'can_view_transactions', 'can_edit_patients', 'is_active']:
        if field in data:
            setattr(emp, field, bool(data[field]))
    
    db.session.commit()
    return jsonify({'success': True, 'data': emp.to_dict()})


@app.route('/api/employees/<int:eid>', methods=['DELETE'])
def api_delete_employee(eid):
    if flask_session.get('user_role') != 'admin':
        return jsonify({'error': 'forbidden', 'message': '仅管理员可执行此操作'}), 403
    emp = Employee.query.get_or_404(eid)
    db.session.delete(emp)
    db.session.commit()
    return jsonify({'success': True})


# =================== 主页 / Dashboard ===================

@app.route('/')
def index():
    return no_cache(render_template('base.html'))


@app.route('/api/dashboard')
def api_dashboard():
    today = date.today()
    month_start = date(today.year, today.month, 1)

    # 今日预约数
    today_appointments = Appointment.query.filter(
        Appointment.scheduled_date == today
    ).count()

    # 今日到访数
    today_visited = Appointment.query.filter(
        Appointment.scheduled_date == today,
        Appointment.actual_visit.isnot(None),
        Appointment.actual_visit != '',
        Appointment.actual_visit != '未到访'
    ).count()

    # 本月成交额
    monthly_revenue = db.session.query(func.sum(Transaction.performance_amount)).filter(
        Transaction.payment_date >= month_start
    ).scalar() or 0

    # 本月成交数
    monthly_deals = Transaction.query.filter(
        Transaction.payment_date >= month_start
    ).count()

    # 本月预约数
    monthly_appointments = Appointment.query.filter(
        Appointment.scheduled_date >= month_start
    ).count()

    # 本月新增患者
    monthly_patients = Patient.query.filter(
        Patient.created_at >= datetime.combine(month_start, datetime.min.time())
    ).count()

    # 总患者数
    total_patients = Patient.query.count()

    # 待跟进（最近30天未成交的初诊）
    thirty_days_ago = today - timedelta(days=30)
    pending_followups = Appointment.query.filter(
        Appointment.scheduled_date >= thirty_days_ago,
        Appointment.scheduled_date <= today,
        or_(
            Appointment.actual_visit == '未成交',
            Appointment.actual_visit == '',
            Appointment.actual_visit.is_(None)
        ),
        Appointment.visit_type == '初诊'
    ).count()

    # 本月业绩趋势（按天）
    revenue_trend = []
    for i in range(min(31, today.day)):
        d = date(today.year, today.month, i + 1)
        day_total = db.session.query(func.sum(Transaction.performance_amount)).filter(
            Transaction.payment_date == d
        ).scalar() or 0
        revenue_trend.append({'date': d.strftime('%m-%d'), 'amount': float(day_total)})

    # 咨询师排名（本月）
    consultant_ranking = db.session.query(
        Transaction.consultant,
        func.sum(Transaction.performance_amount).label('total'),
        func.count(Transaction.id).label('count')
    ).filter(
        Transaction.payment_date >= month_start,
        Transaction.consultant.isnot(None),
        Transaction.consultant != ''
    ).group_by(Transaction.consultant).order_by(func.sum(Transaction.performance_amount).desc()).all()

    # 最近预约
    recent_appointments = Appointment.query.filter(
        Appointment.scheduled_date >= today
    ).order_by(Appointment.scheduled_date).limit(8).all()

    # 最近成交
    recent_transactions = Transaction.query.filter(
        Transaction.payment_date.isnot(None)
    ).order_by(Transaction.payment_date.desc()).limit(8).all()

    return jsonify({
        'today_appointments': today_appointments,
        'today_visited': today_visited,
        'monthly_revenue': float(monthly_revenue),
        'monthly_deals': monthly_deals,
        'monthly_appointments': monthly_appointments,
        'monthly_patients': monthly_patients,
        'total_patients': total_patients,
        'pending_followups': pending_followups,
        'revenue_trend': revenue_trend,
        'consultant_ranking': [
            {'name': r[0] or '未知', 'total': float(r[1] or 0), 'count': r[2]}
            for r in consultant_ranking
        ],
        'recent_appointments': [a.to_dict() for a in recent_appointments],
        'recent_transactions': [t.to_dict() for t in recent_transactions],
    })


# =================== 患者管理 ===================

@app.route('/patients')
def patients_page():
    return no_cache(render_template('base.html'))


@app.route('/api/patients')
def api_patients():
    page = request.args.get('page', 1, type=int)
    per_page = request.args.get('per_page', 20, type=int)
    search = request.args.get('search', '').strip()

    query = Patient.query
    if search:
        query = query.filter(
            or_(
                Patient.name.contains(search),
                Patient.phone.contains(search),
                Patient.internal_id.contains(search),
                Patient.online_consultant.contains(search)
            )
        )

    total = query.count()
    patients = query.order_by(Patient.created_at.desc()).paginate(
        page=page, per_page=per_page, error_out=False
    )

    return jsonify({
        'total': total,
        'page': page,
        'per_page': per_page,
        'pages': (total + per_page - 1) // per_page,
        'data': [p.to_dict() for p in patients.items]
    })


@app.route('/api/patients/<int:pid>')
def api_patient_detail(pid):
    patient = Patient.query.get_or_404(pid)
    appointments = Appointment.query.filter_by(patient_id=pid).order_by(
        Appointment.scheduled_date.desc()
    ).all()
    transactions = Transaction.query.filter_by(patient_id=pid).order_by(
        Transaction.payment_date.desc()
    ).all()
    return jsonify({
        'patient': patient.to_dict(),
        'appointments': [a.to_dict() for a in appointments],
        'transactions': [t.to_dict() for t in transactions],
    })


@app.route('/api/patients', methods=['POST'])
def api_create_patient():
    data = request.get_json()
    phone = normalize_phone(data.get('phone', '')) or None

    # 按电话号码查找重复患者
    existing = None
    if phone:
        existing = Patient.query.filter_by(phone=phone).first()
    if not existing:
        internal_id = data.get('internal_id', '').strip()
        if internal_id:
            existing = Patient.query.filter_by(internal_id=internal_id).first()

    if existing:
        # 合并：更新现有患者的信息
        for field in ['name', 'gender', 'tooth_count', 'condition_desc',
                      'source', 'source_channel', 'online_consultant']:
            val = data.get(field)
            if val:
                setattr(existing, field, val)
        if 'age' in data and data['age']:
            existing.age = int(data['age'])
        existing.updated_at = datetime.now()
        db.session.commit()
        sse_broadcast('patient_updated', existing.to_dict())
        return jsonify({'success': True, 'merged': True, 'data': existing.to_dict()})

    # 自动生成内部编号
    internal_id = data.get('internal_id', '').strip()
    if not internal_id:
        last = Patient.query.order_by(Patient.id.desc()).first()
        next_num = (last.id + 1) if last else 1
        internal_id = f'MEYA-{next_num:04d}'

    p = Patient(
        internal_id=internal_id,
        name=data.get('name', ''),
        phone=phone,
        gender=data.get('gender', ''),
        age=data.get('age', 0, type=int),
        tooth_count=data.get('tooth_count', ''),
        condition_desc=data.get('condition_desc', ''),
        source=data.get('source', ''),
        source_channel=data.get('source_channel', ''),
        online_consultant=data.get('online_consultant', ''),
    )
    db.session.add(p)
    db.session.commit()
    sse_broadcast('patient_created', p.to_dict())
    return jsonify({'success': True, 'merged': False, 'data': p.to_dict()})


@app.route('/api/patients/<int:pid>', methods=['PUT'])
def api_update_patient(pid):
    p = Patient.query.get_or_404(pid)
    data = request.get_json()
    for field in ['name', 'phone', 'gender', 'tooth_count', 'condition_desc',
                  'source', 'source_channel', 'online_consultant', 'internal_id']:
        if field in data:
            setattr(p, field, data[field])
    if 'age' in data and data['age']:
        p.age = int(data['age'])
    p.updated_at = datetime.now()
    db.session.commit()
    sse_broadcast('patient_updated', p.to_dict())
    return jsonify({'success': True, 'data': p.to_dict()})


@app.route('/api/patients/<int:pid>', methods=['DELETE'])
def api_delete_patient(pid):
    p = Patient.query.get_or_404(pid)
    db.session.delete(p)
    db.session.commit()
    sse_broadcast('patient_deleted', {'id': pid})
    return jsonify({'success': True})


# =================== 预约管理 ===================

@app.route('/appointments')
def appointments_page():
    return no_cache(render_template('base.html'))


@app.route('/api/appointments')
def api_appointments():
    perms = flask_session.get('permissions', {})
    if not perms.get('is_admin') and not perms.get('can_view_appointments'):
        return jsonify({'error': 'forbidden', 'message': '权限不足'}), 403
    page = request.args.get('page', 1, type=int)
    per_page = request.args.get('per_page', 20, type=int)
    status = request.args.get('status', '')
    search = request.args.get('search', '').strip()
    month = request.args.get('month', '')

    query = Appointment.query

    if search:
        query = query.join(Patient).filter(
            or_(
                Patient.name.contains(search),
                Patient.phone.contains(search),
                Patient.internal_id.contains(search),
                Appointment.inviter.contains(search),
                Appointment.phone.contains(search),
            )
        )

    if status == 'today':
        query = query.filter(Appointment.scheduled_date == date.today())
    elif status == 'upcoming':
        query = query.filter(Appointment.scheduled_date >= date.today())
    elif status == 'past':
        query = query.filter(Appointment.scheduled_date < date.today())
    elif status == 'visited':
        query = query.filter(
            Appointment.actual_visit.isnot(None),
            Appointment.actual_visit != '',
            Appointment.actual_visit != '未到访'
        )
    elif status == 'noshow':
        query = query.filter(Appointment.actual_visit == '未到访')
    elif status == 'nodata':
        query = query.filter(
            or_(
                Appointment.actual_visit.is_(None),
                Appointment.actual_visit == ''
            )
        )
    elif status == 'deals':
        query = query.filter(
            Appointment.actual_visit.in_(['成交', '定金', '定金转成交', '未成交转成交'])
        )

    if month:
        try:
            y, m = month.split('-')
            query = query.filter(
                extract('year', Appointment.scheduled_date) == int(y),
                extract('month', Appointment.scheduled_date) == int(m)
            )
        except:
            pass

    total = query.count()
    appointments = query.order_by(Appointment.scheduled_date.desc()).paginate(
        page=page, per_page=per_page, error_out=False
    )

    return jsonify({
        'total': total,
        'page': page,
        'per_page': per_page,
        'pages': (total + per_page - 1) // per_page,
        'data': [a.to_dict() for a in appointments.items]
    })


@app.route('/api/appointments', methods=['POST'])
def api_create_appointment():
    data = request.get_json()
    patient_id = data.get('patient_id')

    # 如果没传 patient_id，按电话/内部编号/姓名查找或创建患者
    if not patient_id:
        phone = normalize_phone(data.get('phone', '')) or None
        internal_id = data.get('internal_id', '').strip()
        name = data.get('patient_name', '').strip()

        patient = None
        if phone:
            patient = Patient.query.filter_by(phone=phone).first()
        if not patient and internal_id:
            patient = Patient.query.filter_by(internal_id=internal_id).first()
        if not patient and name:
            patient = Patient.query.filter_by(name=name).first()

        if not patient:
            if not phone and not internal_id and not name:
                return jsonify({'success': False, 'error': '请提供患者信息（姓名/电话/编号）'}), 400
            last = Patient.query.order_by(Patient.id.desc()).first()
            next_num = (last.id + 1) if last else 1
            internal_id = internal_id or f'MEYA-{next_num:04d}'
            patient = Patient(internal_id=internal_id, name=name, phone=phone)
            db.session.add(patient)
            db.session.flush()
        patient_id = patient.id

    a = Appointment(
        patient_id=patient_id,
        scheduled_date=_parse_date(data.get('scheduled_date')),
        scheduled_time=data.get('scheduled_time', ''),
        inviter=data.get('inviter', ''),
        inviter_2=data.get('inviter_2', ''),
        phone=normalize_phone(data.get('phone', '')),
        visit_type=data.get('visit_type', ''),
        actual_visit=data.get('actual_visit', ''),
        consultation_notes=data.get('consultation_notes', ''),
        followup_24h=data.get('followup_24h', ''),
        followup_7d=data.get('followup_7d', ''),
        followup_15d=data.get('followup_15d', ''),
        followup_30d=data.get('followup_30d', ''),
        need_collab=data.get('need_collab', ''),
        invalid_reason=data.get('invalid_reason', ''),
        has_continue=data.get('has_continue', ''),
        registered_date=_parse_date(data.get('registered_date')),
    )
    db.session.add(a)
    db.session.commit()
    sse_broadcast('appointment_created', a.to_dict())
    return jsonify({'success': True, 'data': a.to_dict()})


@app.route('/api/appointments/<int:aid>', methods=['PUT'])
def api_update_appointment(aid):
    a = Appointment.query.get_or_404(aid)
    data = request.get_json()
    for field in ['scheduled_time', 'inviter', 'inviter_2', 'phone', 'visit_type',
                  'actual_visit', 'consultation_notes', 'followup_24h', 'followup_7d',
                  'followup_15d', 'followup_30d', 'need_collab', 'invalid_reason',
                  'has_continue']:
        if field in data:
            setattr(a, field, data[field])
    if 'scheduled_date' in data:
        a.scheduled_date = _parse_date(data['scheduled_date'])
    if 'registered_date' in data:
        a.registered_date = _parse_date(data['registered_date'])
    if 'patient_id' in data and data['patient_id']:
        a.patient_id = data['patient_id']
    db.session.commit()
    sse_broadcast('appointment_updated', a.to_dict())
    return jsonify({'success': True, 'data': a.to_dict()})


@app.route('/api/appointments/<int:aid>', methods=['DELETE'])
def api_delete_appointment(aid):
    a = Appointment.query.get_or_404(aid)
    db.session.delete(a)
    db.session.commit()
    sse_broadcast('appointment_deleted', {'id': aid})
    return jsonify({'success': True})


# =================== 成交管理 ===================

@app.route('/transactions')
def transactions_page():
    return no_cache(render_template('base.html'))


@app.route('/api/transactions')
def api_transactions():
    perms = flask_session.get('permissions', {})
    if not perms.get('is_admin') and not perms.get('can_view_transactions'):
        return jsonify({'error': 'forbidden', 'message': '权限不足'}), 403
    page = request.args.get('page', 1, type=int)
    per_page = request.args.get('per_page', 20, type=int)
    search = request.args.get('search', '').strip()
    month = request.args.get('month', '')

    query = Transaction.query

    if search:
        query = query.join(Patient).filter(
            or_(
                Patient.name.contains(search),
                Transaction.consultant.contains(search),
                Transaction.plan_type.contains(search),
            )
        )

    if month:
        try:
            y, m = month.split('-')
            query = query.filter(
                extract('year', Transaction.payment_date) == int(y),
                extract('month', Transaction.payment_date) == int(m)
            )
        except:
            pass

    total = query.count()
    transactions = query.order_by(Transaction.payment_date.desc()).paginate(
        page=page, per_page=per_page, error_out=False
    )

    return jsonify({
        'total': total,
        'page': page,
        'per_page': per_page,
        'pages': (total + per_page - 1) // per_page,
        'data': [t.to_dict() for t in transactions.items]
    })


@app.route('/api/transactions', methods=['POST'])
def api_create_transaction():
    data = request.get_json()
    patient_id = data.get('patient_id')

    if not patient_id:
        phone = normalize_phone(data.get('phone', '')) or None
        internal_id = data.get('internal_id', '').strip()
        name = data.get('patient_name', '').strip()

        # 优先按电话号码查找，其次内部编号，最后姓名
        patient = None
        if phone:
            patient = Patient.query.filter_by(phone=phone).first()
        if not patient and internal_id:
            patient = Patient.query.filter_by(internal_id=internal_id).first()
        if not patient and name:
            patient = Patient.query.filter_by(name=name).first()

        if not patient:
            # 无任何标识信息，无法创建患者
            if not phone and not internal_id and not name:
                return jsonify({'success': False, 'error': '请提供患者信息（姓名/电话/编号）'}), 400
            # 自动生成内部编号
            last = Patient.query.order_by(Patient.id.desc()).first()
            next_num = (last.id + 1) if last else 1
            internal_id = internal_id or f'MEYA-{next_num:04d}'
            patient = Patient(
                internal_id=internal_id,
                name=name,
                phone=phone,
            )
            db.session.add(patient)
            db.session.flush()
        patient_id = patient.id

    t = Transaction(
        patient_id=patient_id,
        plan_type=data.get('plan_type', ''),
        plan_detail=data.get('plan_detail', ''),
        brand=data.get('brand', ''),
        tooth_positions=data.get('tooth_positions', ''),
        tooth_count=data.get('tooth_count', 0, type=int),
        bone_graft=data.get('bone_graft', ''),
        bone_membrane=data.get('bone_membrane', ''),
        sinus_lift=data.get('sinus_lift', ''),
        deposit_amount=float(data.get('deposit_amount', 0) or 0),
        deposit_contract=float(data.get('deposit_contract', 0) or 0),
        deposit_date=_parse_date(data.get('deposit_date')),
        deposit_return_date=_parse_date(data.get('deposit_return_date')),
        performance_amount=float(data.get('performance_amount', 0) or 0),
        paid_amount=float(data.get('paid_amount', 0) or 0),
        payment_date=_parse_date(data.get('payment_date')),
        total_amount=float(data.get('total_amount', 0) or 0),
        payment=float(data.get('payment', 0) or 0),
        payment_date_2=_parse_date(data.get('payment_date_2')),
        supplement_amount=float(data.get('supplement_amount', 0) or 0),
        supplement_date=_parse_date(data.get('supplement_date')),
        debt_amount=float(data.get('debt_amount', 0) or 0),
        supplement_records=data.get('supplement_records', ''),
        treatment_date=_parse_date(data.get('treatment_date')),
        treatment_doctor=data.get('treatment_doctor', ''),
        visit_outcome=data.get('visit_outcome', ''),
        consultant=data.get('consultant', ''),
    )
    db.session.add(t)
    db.session.commit()
    sse_broadcast('transaction_created', t.to_dict())
    return jsonify({'success': True, 'data': t.to_dict()})


@app.route('/api/transactions/<int:tid>', methods=['PUT'])
def api_update_transaction(tid):
    t = Transaction.query.get_or_404(tid)
    data = request.get_json()
    for field in ['plan_type', 'plan_detail', 'brand', 'tooth_positions',
                  'bone_graft', 'bone_membrane', 'sinus_lift', 'supplement_records',
                  'treatment_doctor', 'visit_outcome', 'consultant']:
        if field in data:
            setattr(t, field, data[field])
    for num_field in ['tooth_count', 'deposit_amount', 'deposit_contract',
                      'performance_amount', 'paid_amount', 'total_amount', 'payment',
                      'supplement_amount', 'debt_amount']:
        if num_field in data:
            setattr(t, num_field, float(data[num_field] or 0))
    for date_field in ['deposit_date', 'deposit_return_date', 'payment_date',
                       'payment_date_2', 'supplement_date', 'treatment_date']:
        if date_field in data:
            setattr(t, date_field, _parse_date(data[date_field]))
    if 'patient_id' in data and data['patient_id']:
        t.patient_id = data['patient_id']
    db.session.commit()
    sse_broadcast('transaction_updated', t.to_dict())
    return jsonify({'success': True, 'data': t.to_dict()})


@app.route('/api/transactions/<int:tid>', methods=['DELETE'])
def api_delete_transaction(tid):
    t = Transaction.query.get_or_404(tid)
    db.session.delete(t)
    db.session.commit()
    sse_broadcast('transaction_deleted', {'id': tid})
    return jsonify({'success': True})


# =================== 报表 ===================

@app.route('/reports')
def reports_page():
    return no_cache(render_template('base.html'))


@app.route('/api/reports/overview')
def api_reports_overview():
    perms = flask_session.get('permissions', {})
    if not perms.get('is_admin') and not perms.get('can_view_transactions'):
        return jsonify({'error': 'forbidden', 'message': '权限不足'}), 403
    year = request.args.get('year', date.today().year, type=int)

    # 月度业绩
    monthly_revenue = []
    for m in range(1, 13):
        total = db.session.query(func.sum(Transaction.performance_amount)).filter(
            extract('year', Transaction.payment_date) == year,
            extract('month', Transaction.payment_date) == m
        ).scalar() or 0
        monthly_revenue.append({'month': f'{m}月', 'amount': float(total)})

    # 咨询师年度排名
    consultant_ranking = db.session.query(
        Transaction.consultant,
        func.sum(Transaction.performance_amount).label('total'),
        func.count(Transaction.id).label('count')
    ).filter(
        extract('year', Transaction.payment_date) == year,
        Transaction.consultant.isnot(None),
        Transaction.consultant != ''
    ).group_by(Transaction.consultant).order_by(
        func.sum(Transaction.performance_amount).desc()
    ).all()

    # 来源分析
    source_analysis = db.session.query(
        Patient.source,
        func.count(Patient.id)
    ).filter(
        Patient.source.isnot(None),
        Patient.source != ''
    ).group_by(Patient.source).all()

    # 到访情况分析
    visit_analysis = db.session.query(
        Appointment.actual_visit,
        func.count(Appointment.id)
    ).filter(
        Appointment.actual_visit.isnot(None),
        Appointment.actual_visit != ''
    ).group_by(Appointment.actual_visit).all()

    return jsonify({
        'monthly_revenue': monthly_revenue,
        'consultant_ranking': [
            {'name': r[0], 'total': float(r[1] or 0), 'count': r[2]}
            for r in consultant_ranking
        ],
        'source_analysis': [
            {'name': r[0], 'count': r[1]} for r in source_analysis
        ],
        'visit_analysis': [
            {'name': r[0], 'count': r[1]} for r in visit_analysis
        ],
    })


# =================== 数据导入 ===================

@app.route('/import')
def import_page():
    return no_cache(render_template('base.html'))


@app.route('/api/import/upload', methods=['POST'])
def api_import_upload():
    """上传 Excel 文件到临时目录"""
    if 'file' not in request.files:
        return jsonify({'success': False, 'error': '未选择文件'}), 400
    f = request.files['file']
    if f.filename == '':
        return jsonify({'success': False, 'error': '文件名为空'}), 400
    upload_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'uploads')
    os.makedirs(upload_dir, exist_ok=True)
    ext = os.path.splitext(f.filename)[1]
    save_name = str(uuid.uuid4()) + ext
    save_path = os.path.join(upload_dir, save_name)
    f.save(save_path)
    return jsonify({'success': True, 'path': save_path, 'filename': f.filename})


@app.route('/api/import/preview', methods=['POST'])
def api_import_preview():
    """预览 Excel 文件内容"""
    import openpyxl
    data = request.get_json()
    path = data.get('path', '')
    sheet_name = data.get('sheet', '')

    if not os.path.exists(path):
        return jsonify({'success': False, 'error': '文件不存在'}), 400

    try:
        wb = openpyxl.load_workbook(path, data_only=True)
        sheets = wb.sheetnames

        ws = wb[sheet_name] if sheet_name else wb.active
        rows = []
        for i, row in enumerate(ws.iter_rows(min_row=1, max_row=min(20, ws.max_row), values_only=True)):
            rows.append([str(c) if c is not None else '' for c in row])

        return jsonify({
            'success': True,
            'sheets': sheets,
            'current_sheet': ws.title,
            'total_rows': ws.max_row,
            'total_cols': ws.max_column,
            'preview': rows,
        })
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/import/execute', methods=['POST'])
def api_import_execute():
    """执行导入"""
    import openpyxl
    from datetime import datetime as dt

    data = request.get_json()
    path = data.get('path', '')
    sheet_name = data.get('sheet', '')
    table_type = data.get('table_type', 'appointment')  # appointment or performance

    if not os.path.exists(path):
        return jsonify({'success': False, 'error': '文件不存在'}), 400

    try:
        wb = openpyxl.load_workbook(path, data_only=True)
        ws = wb[sheet_name] if sheet_name else wb.active

        headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
        imported = 0
        errors = []

        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            try:
                vals = [str(c) if c is not None else '' for c in row]
                row_dict = dict(zip(headers, vals))

                if table_type == 'appointment':
                    _import_appointment_row(row_dict)
                else:
                    _import_performance_row(row_dict)

                imported += 1
            except Exception as e:
                errors.append(f'行{row_idx}: {str(e)}')
                continue

        db.session.commit()
        return jsonify({
            'success': True,
            'imported': imported,
            'errors': errors[:20],
        })
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


# =================== SSE 实时推送 ===================

@app.route('/api/events')
def sse_events():
    from queue import Queue
    q = Queue()
    sse_clients.append(q)

    def stream():
        try:
            yield 'event: connected\ndata: {}\n\n'
            while True:
                msg = q.get()
                yield msg
        except GeneratorExit:
            if q in sse_clients:
                sse_clients.remove(q)

    return Response(stream(), mimetype='text/event-stream')


# =================== Debug ===================

@app.route('/api/debug/employees')
def api_debug_employees():
    """Debug: check employees table."""
    import traceback, sys
    try:
        from sqlalchemy import inspect
        inspector = inspect(db.engine)
        tables = inspector.get_table_names()
        
        # Try to fix tables
        db.create_all()
        
        err_msg = None
        emp_count = -1
        try:
            emp_count = Employee.query.count()
        except Exception as e2:
            err_msg = str(e2)
        
        return jsonify({
            'tables': tables,
            'employee_table_exists': 'employees' in tables,
            'employee_count': emp_count,
            'error': err_msg,
            'permissions': flask_session.get('permissions', {}),
            'role': flask_session.get('user_role', ''),
        })
    except Exception as e:
        return jsonify({'error': str(e), 'traceback': traceback.format_exc()}), 500


# =================== 搜索补全 ===================

@app.route('/api/search/patients')
def api_search_patients():
    q = request.args.get('q', '').strip()
    if not q:
        return jsonify([])
    results = Patient.query.filter(
        or_(
            Patient.name.contains(q),
            Patient.phone.contains(q),
            Patient.internal_id.contains(q),
        )
    ).limit(10).all()
    return jsonify([{
        'id': p.id, 'name': p.name, 'phone': p.phone or '',
        'internal_id': p.internal_id or ''
    } for p in results])


# =================== 工具函数 ===================

def _parse_date(s):
    if not s or s == '':
        return None
    if isinstance(s, date) and not isinstance(s, datetime):
        return s
    if isinstance(s, datetime):
        return s.date()
    s = str(s).strip()
    for fmt in ['%Y-%m-%d', '%Y年%m月%d日', '%Y/%m/%d', '%m月%d日', '%d/%m/%Y']:
        try:
            return dt.strptime(s, fmt).date()
        except ValueError:
            continue
    # Handle formats like "5.31"
    import re
    m = re.match(r'(\d+)\.(\d+)', s)
    if m:
        try:
            return date(2026, int(m.group(1)), int(m.group(2)))
        except:
            pass
    return None


def _import_appointment_row(row):
    """导入预约总表的一行数据"""
    patient_name = row.get('患者姓名', '').strip().replace('\xa0', ' ').strip()
    if not patient_name or patient_name == 'None':
        return

    phone = (str(row.get('电话号码', '') or '').strip().replace('\xa0', '') or None)
    if phone == 'None' or phone == '':
        phone = ''

    # 查找或创建患者: 优先按电话查找，其次姓名
    patient = None
    if phone:
        patient = Patient.query.filter_by(phone=phone).first()
    if not patient and patient_name:
        patient = Patient.query.filter_by(name=patient_name).first()
    if not patient:
        last = Patient.query.order_by(Patient.id.desc()).first()
        next_num = (last.id + 1) if last else 1
        internal_id = f'MEYA-{next_num:04d}'
        patient = Patient(
            internal_id=internal_id,
            name=patient_name,
            phone=phone,
            tooth_count=row.get('缺牙颗数', ''),
            source=row.get('客户来源', ''),
            source_channel=row.get('线索来源渠道', ''),
            online_consultant=row.get('邀约人', ''),
        )
        db.session.add(patient)
        db.session.flush()

    a = Appointment(
        patient_id=patient.id,
        scheduled_date=_parse_date(row.get('预计到访日期', '')),
        scheduled_time=row.get('预计具体时间', ''),
        inviter=row.get('邀约人', ''),
        inviter_2=row.get('邀约人 2', ''),
        phone=phone,
        visit_type=row.get('到访类型', ''),
        actual_visit=row.get('实际到访情况', ''),
        consultation_notes=row.get('咨询及跟进情况，日期+内容', ''),
        registered_date=_parse_date(row.get('登记日期', '')),
    )
    db.session.add(a)
    db.session.flush()

    # 如果有成交信息，也创建交易记录
    contract_amt = _parse_float(row.get('定金合同额', '0'))
    deposit_amt = _parse_float(row.get('定金额', '0'))
    performance_amt = _parse_float(row.get('成交额', '0'))
    paid_amt = _parse_float(row.get('支付额', '0'))
    supplement_amt = _parse_float(row.get('补费额', '0'))

    if contract_amt > 0 or performance_amt > 0:
        t = Transaction(
            patient_id=patient.id,
            plan_type=row.get('方案类型', ''),
            plan_detail=row.get('方案（品牌、颗数、牙位）', ''),
            brand=row.get('品牌', ''),
            tooth_positions=row.get('牙位', ''),
            tooth_count=_parse_int(row.get('颗数', '0')),
            bone_graft=row.get('骨粉', ''),
            bone_membrane=row.get('骨膜', ''),
            sinus_lift=row.get('内外提', ''),
            deposit_amount=deposit_amt,
            deposit_contract=contract_amt,
            deposit_date=_parse_date(row.get('定金日期', '')),
            deposit_return_date=_parse_date(row.get('定金回款日期', '')),
            performance_amount=performance_amt,
            paid_amount=paid_amt,
            payment_date=_parse_date(row.get('成交日期', '')),
            supplement_amount=supplement_amt,
            supplement_date=_parse_date(row.get('补费日期', '')),
            debt_amount=_parse_float(row.get('欠款额', '0')),
            supplement_records=row.get('补费记录（日期+金额）', ''),
            treatment_date=_parse_date(row.get('治疗日期', '')),
            treatment_doctor=row.get('治疗医生', ''),
            consultant=row.get('谈单人', ''),
            visit_outcome=row.get('成交类型', ''),
        )
        db.session.add(t)


def _import_performance_row(row):
    """导入接诊业绩表的一行数据"""
    patient_name = row.get('患者名字', '').strip().replace('\xa0', ' ').strip()
    if not patient_name or patient_name == 'None':
        return

    phone = (str(row.get('电话号码', '') or '').strip().replace('\xa0', '') or None)
    if phone == 'None' or phone == '':
        phone = ''

    # 按电话 > 姓名查找已有患者
    patient = None
    if phone:
        patient = Patient.query.filter_by(phone=phone).first()
    if not patient and patient_name:
        patient = Patient.query.filter_by(name=patient_name).first()
    if not patient:
        last = Patient.query.order_by(Patient.id.desc()).first()
        next_num = (last.id + 1) if last else 1
        internal_id = f'MEYA-{next_num:04d}'
        patient = Patient(
            internal_id=internal_id,
            name=patient_name,
            phone=phone,
            online_consultant=row.get('网咨', ''),
            condition_desc=row.get('患者情况', ''),
        )
        db.session.add(patient)
        db.session.flush()

    # 创建预约记录
    a = Appointment(
        patient_id=patient.id,
        scheduled_date=_parse_date(row.get('日期', '')),
        visit_type=row.get('初/复诊', ''),
        actual_visit=row.get('实际到访情况', ''),
        followup_24h=row.get('跟进记录(24h)', ''),
        followup_7d=row.get('跟进记录(7天)', ''),
        followup_15d=row.get('跟进记录(15天)', ''),
        followup_30d=row.get('跟进记录(30天)', ''),
        need_collab=row.get('是否需要协同跟进', ''),
        invalid_reason=row.get('无效原因', ''),
        has_continue=row.get('是否有续种', ''),
    )
    db.session.add(a)

    # 创建交易记录
    perf_amt = _parse_float(row.get('业绩金额(半款)', '0'))
    paid_amt = _parse_float(row.get('已付金额(半款)', '0'))
    deposit_amt = _parse_float(row.get('定金金额', '0'))
    contract_amt = _parse_float(row.get('定金合同金额', '0'))

    t = Transaction(
        patient_id=patient.id,
        plan_type=row.get('方案类型', ''),
        performance_amount=perf_amt,
        paid_amount=paid_amt,
        payment_date=_parse_date(row.get('交款日期', '')),
        deposit_amount=deposit_amt,
        deposit_contract=contract_amt,
        deposit_return_date=_parse_date(row.get('定金回款日期', '')),
        consultant=row.get('咨询', ''),
    )
    db.session.add(t)


# =================== Lazy Initialization ===================
# Instead of running at module load (which blocks gunicorn startup),
# we initialize on the first request. This eliminates 502 errors.

_app_initialized = False
_init_error = None

def ensure_initialized():
    """Run on first request. Creates tables + imports if needed."""
    global _app_initialized, _init_error
    if _app_initialized:
        return True
    
    try:
        from sqlalchemy import text as sa_text
        
        # Try WAL mode; non-fatal if locked (database works without it)
        try:
            db.session.execute(sa_text('PRAGMA journal_mode=WAL'))
            db.session.execute(sa_text('PRAGMA synchronous=NORMAL'))
            db.session.commit()
        except Exception:
            db.session.rollback()
        
        # Create employees table if not exists
        try:
            db.session.execute(sa_text('''
                CREATE TABLE IF NOT EXISTS employees (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    username VARCHAR(50) NOT NULL UNIQUE,
                    password VARCHAR(200) NOT NULL,
                    display_name VARCHAR(50) DEFAULT '',
                    role VARCHAR(20) DEFAULT 'staff',
                    can_view_appointments BOOLEAN DEFAULT 1,
                    can_view_transactions BOOLEAN DEFAULT 1,
                    can_edit_patients BOOLEAN DEFAULT 1,
                    is_active BOOLEAN DEFAULT 1,
                    created_at TIMESTAMP,
                    updated_at TIMESTAMP
                )
            '''))
            db.session.commit()
        except:
            pass
        
        db.create_all()

        # ===== Migration: add internal_id column if missing in existing DB =====
        from sqlalchemy import inspect as sa_inspect
        inspector = sa_inspect(db.engine)
        cols = [c['name'] for c in inspector.get_columns('patients')]
        if 'internal_id' not in cols:
            try:
                db.session.execute(sa_text('ALTER TABLE patients ADD COLUMN internal_id VARCHAR(20)'))
                db.session.commit()
                print('📌 Added internal_id column')
            except Exception:
                db.session.rollback()

        # Generate internal_id for existing patients
        missing = Patient.query.filter(Patient.internal_id.is_(None)).all()
        for p in missing:
            p.internal_id = f'MEYA-{p.id:04d}'
        if missing:
            db.session.commit()
            print(f'📌 Generated internal_id for {len(missing)} patients')

        # Handle duplicate phone numbers (remove unique constraint is complex,
        # instead just dedupe on the fly)
        dupe_phones = db.session.query(
            Patient.phone, db.func.count(Patient.id)
        ).filter(
            Patient.phone != '', Patient.phone.isnot(None)
        ).group_by(Patient.phone).having(db.func.count(Patient.id) > 1).all()
        for phone, cnt in dupe_phones:
            patients = Patient.query.filter_by(phone=phone).order_by(Patient.id).all()
            keep = patients[0]
            for dup in patients[1:]:
                # Merge appointments and transactions to the kept patient
                Appointment.query.filter_by(patient_id=dup.id).update({'patient_id': keep.id})
                Transaction.query.filter_by(patient_id=dup.id).update({'patient_id': keep.id})
                db.session.delete(dup)
            print(f'📌 Merged {cnt-1} duplicate(s) for phone {phone}')
        if dupe_phones:
            db.session.commit()
        
        # Import data if empty
        if Patient.query.count() == 0:
            print('📥 Auto-importing Excel data...')
            import import_data
            import_data.import_from_files(app)
            print(f'✅ Imported {Patient.query.count()} patients')
        
        _app_initialized = True
        _init_error = None
        return True
    except Exception as e:
        import traceback
        _init_error = str(e)
        print(f'⚠️ Init failed: {e}')
        traceback.print_exc()
        return False


@app.before_request
def check_initialized():
    """Ensure app is initialized before handling any request."""
    if request.path in ['/api/health', '/health']:
        return None
    if request.path.startswith('/static/'):
        return None
    if not _app_initialized:
        with app.app_context():
            ok = ensure_initialized()
        if not ok:
            return jsonify({
                'error': 'initializing',
                'message': f'服务启动中: {_init_error}'
            }), 503


@app.route('/api/health')
def api_health():
    ok = _app_initialized
    return jsonify({
        'status': 'ok' if ok else 'starting',
        'initialized': ok,
        'error': _init_error,
        'patients': Patient.query.count() if ok else -1,
        'tables': [] if ok else None,
    })


@app.route('/health')
def health_page():
    return api_health()


@app.route('/api/fix/phones')
def api_fix_phones():
    """One-time: add 60 prefix to all patient phone numbers."""
    if flask_session.get('user_role') != 'admin':
        return jsonify({'error': 'forbidden'}), 403
    updated = 0
    for p in Patient.query.filter(Patient.phone != '', Patient.phone.isnot(None)).all():
        phone = p.phone.strip()
        if phone and not phone.startswith('60'):
            p.phone = '60' + phone
            updated += 1
    db.session.commit()
    return jsonify({'success': True, 'updated': updated})


def _parse_float(s):
    if not s or s == '' or s == 'None':
        return 0.0
    s = str(s).replace('RM', '').replace('RM', '').replace(',', '').replace(' ', '')
    try:
        return float(s)
    except:
        return 0.0


def _parse_int(s):
    if not s or s == '' or s == 'None':
        return 0
    try:
        return int(float(str(s).strip()))
    except:
        return 0


# Initialization happens on first request (see check_initialized above)

if __name__ == '__main__':

    port = int(os.environ.get('PORT', 5000))
    debug = os.environ.get('FLASK_ENV') == 'development'
    host = '0.0.0.0' if os.environ.get('RENDER') else '127.0.0.1'
    app.run(host=host, port=port, debug=debug)
