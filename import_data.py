"""Import all Excel data into the database.
Avoids circular imports by importing models directly.
"""
import sys
import os
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

from datetime import datetime, date
import openpyxl
import re

# Direct model imports (not from app, so no circular dependency)
from models import db, Patient, Appointment, Transaction

# Data files - look relative to script first, then fallback
_script_dir = os.path.dirname(os.path.abspath(__file__))
DATA_DIR = os.path.join(_script_dir, 'data') if os.path.isdir(os.path.join(_script_dir, 'data')) \
    else '/Users/meiya/.openclaw/workspace/数据表格'


def parse_date(s):
    if not s or str(s).strip() == '':
        return None
    if isinstance(s, (datetime, date)):
        d = s if isinstance(s, date) else s.date()
        # Excel empty cells become 1900 dates
        if d.year < 2000:
            return None
        return d
    s = str(s).strip().replace('\xa0', ' ')
    
    m = re.match(r'(\d{1,2})[./](\d{1,2})', s)
    if m:
        try:
            m_num, d_num = int(m.group(1)), int(m.group(2))
            if m_num > 12 and d_num <= 12:
                m_num, d_num = d_num, m_num
            return date(2026, m_num, d_num)
        except:
            pass
    
    for fmt in ['%Y-%m-%d', '%Y年%m月%d日', '%Y/%m/%d', '%m月%d日', '%d/%m/%Y']:
        try:
            d = datetime.strptime(s, fmt).date()
            if d.year >= 2000:
                return d
        except:
            continue
    return None


def parse_float(s):
    if s is None or str(s).strip() == '':
        return 0.0
    if isinstance(s, (int, float)):
        return float(s)
    s = str(s).replace('RM', '').replace('RM', '').replace(',', '').replace(' ', '').replace('\xa0', '')
    try:
        return float(s)
    except:
        return 0.0


def parse_int(s):
    if s is None or str(s).strip() == '':
        return 0
    if isinstance(s, (int, float)):
        return int(s)
    try:
        return int(float(str(s).strip()))
    except:
        return 0


def clean_name(name):
    if not name:
        return ''
    return str(name).replace('\xa0', ' ').strip()


# 内部编号计数器，按年月分组
_id_counters = {}

def generate_internal_id(scheduled_date):
    """生成格式: YYYY-MM-NNNN"""
    if scheduled_date:
        if isinstance(scheduled_date, date):
            ym = scheduled_date.strftime('%Y-%m')
        else:
            ym = str(scheduled_date)[:7]
    else:
        ym = date.today().strftime('%Y-%m')
    if ym not in _id_counters:
        _id_counters[ym] = 0
    _id_counters[ym] += 1
    return f'{ym}-{_id_counters[ym]:04d}'


def import_performance(wb, sheet_name):
    ws = wb[sheet_name]
    headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    
    imported = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        vals = [str(v) if v is not None else '' for v in row]
        row_dict = dict(zip(headers, vals))
        
        patient_name = clean_name(row_dict.get('患者名字', ''))
        if not patient_name or patient_name == 'None':
            continue
        
        phone = (str(row_dict.get('电话号码', '') or '').strip() or None)
        if phone == 'None' or not phone:
            phone = None
        patient = None
        if phone:
            patient = Patient.query.filter_by(phone=phone).first()
        if not patient and patient_name:
            patient = Patient.query.filter_by(name=patient_name).first()
        if not patient:
            sd = parse_date(row_dict.get('日期', ''))
            internal_id = generate_internal_id(sd)
            patient = Patient(
                internal_id=internal_id,
                name=patient_name,
                phone=phone,
                online_consultant=clean_name(row_dict.get('网咨', '')),
                condition_desc=row_dict.get('患者情况', ''),
            )
            try:
                db.session.add(patient)
                db.session.flush()
            except Exception:
                db.session.rollback()
                # Duplicate found - find existing patient
                if patient_name:
                    patient = Patient.query.filter_by(name=patient_name).first()
                if not patient and phone:
                    patient = Patient.query.filter_by(phone=phone).first()
                if not patient:
                    print(f'⚠️ Skipping duplicate: {patient_name}')
                    continue
        # 检查重复预约
        sd = parse_date(row_dict.get('日期', ''))
        vt = clean_name(row_dict.get('初/复诊', ''))
        existing_appt = Appointment.query.filter_by(patient_id=patient.id, scheduled_date=sd, visit_type=vt).first()
        if not existing_appt:
            appt = Appointment(
                patient_id=patient.id,
                scheduled_date=sd,
                visit_type=vt,
                actual_visit=clean_name(row_dict.get('实际到访情况', '')),
                followup_24h=row_dict.get('跟进记录(24h)', ''),
                followup_7d=row_dict.get('跟进记录(7天)', ''),
                followup_15d=row_dict.get('跟进记录(15天)', ''),
                followup_30d=row_dict.get('跟进记录(30天)', ''),
                need_collab=clean_name(row_dict.get('是否需要协同跟进', '')),
                invalid_reason=row_dict.get('无效原因', ''),
                has_continue=clean_name(row_dict.get('是否有续种', '')),
            )
            db.session.add(appt)
        
        # 检查重复交易
        pd2 = parse_date(row_dict.get('交款日期', ''))
        pt = clean_name(row_dict.get('方案类型', ''))
        pa = parse_float(row_dict.get('业绩金额(半款)', '0'))
        existing_txn = Transaction.query.filter_by(patient_id=patient.id, payment_date=pd2, plan_type=pt, performance_amount=pa).first()
        if not existing_txn:
            txn = Transaction(
                patient_id=patient.id,
                plan_type=pt,
                performance_amount=pa,
                paid_amount=parse_float(row_dict.get('已付金额(半款)', '0')),
                payment_date=pd2,
                deposit_amount=parse_float(row_dict.get('定金金额', '0')),
                deposit_contract=parse_float(row_dict.get('定金合同金额', '0')),
                deposit_return_date=parse_date(row_dict.get('定金回款日期', '')),
                consultant=clean_name(row_dict.get('咨询', '')),
            )
            db.session.add(txn)
        imported += 1
    
    return imported


def import_appointment(wb, sheet_name):
    ws = wb[sheet_name]
    headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    
    imported = 0
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        vals = [str(v) if v is not None else '' for v in row]
        row_dict = dict(zip(headers, vals))
        
        patient_name = clean_name(row_dict.get('患者姓名', ''))
        if not patient_name or patient_name == 'None':
            continue
        
        phone = str(row_dict.get('电话号码', '') or '').replace('\xa0', '').strip()
        if phone == 'None' or not phone:
            phone = None
        if phone and not phone.startswith('60'):
            phone = '60' + phone
        
        patient = None
        if phone:
            patient = Patient.query.filter_by(phone=phone).first()
        if not patient and patient_name:
            patient = Patient.query.filter_by(name=patient_name).first()
        if not patient:
            sd = parse_date(row_dict.get('预计到访日期', ''))
            internal_id = generate_internal_id(sd)
            patient = Patient(
                internal_id=internal_id,
                name=patient_name,
                phone=phone,
                tooth_count=clean_name(row_dict.get('缺牙颗数', '')),
                source=clean_name(row_dict.get('客户来源', '')),
                source_channel=clean_name(row_dict.get('线索来源渠道', '')),
                online_consultant=clean_name(row_dict.get('邀约人', '')),
            )
            db.session.add(patient)
            db.session.flush()
        
        # 检查重复预约
        appt_sd = parse_date(row_dict.get('预计到访日期', ''))
        appt_vt = clean_name(row_dict.get('到访类型', ''))
        existing_appt = Appointment.query.filter_by(patient_id=patient.id, scheduled_date=appt_sd, visit_type=appt_vt).first()
        if not existing_appt:
            appt = Appointment(
                patient_id=patient.id,
                scheduled_date=appt_sd,
                scheduled_time=clean_name(row_dict.get('预计具体时间', '')),
                inviter=clean_name(row_dict.get('邀约人', '')),
                inviter_2=clean_name(row_dict.get('邀约人 2', '')),
                phone=phone,
                visit_type=appt_vt,
                actual_visit=clean_name(row_dict.get('实际到访情况', '')),
                consultation_notes=row_dict.get('咨询及跟进情况，日期+内容', ''),
                registered_date=parse_date(row_dict.get('登记日期', '')),
            )
            db.session.add(appt)
        
        contract_amt = parse_float(row_dict.get('定金合同额', '0'))
        deposit_amt = parse_float(row_dict.get('定金额', '0'))
        perf_amt = parse_float(row_dict.get('成交额', '0'))
        paid_amt = parse_float(row_dict.get('支付额', '0'))
        supp_amt = parse_float(row_dict.get('补费额', '0'))
        debt_amt = parse_float(row_dict.get('欠款额', '0'))
        
        if contract_amt > 0 or perf_amt > 0 or paid_amt > 0:
            txn_pd = parse_date(row_dict.get('成交日期', ''))
            txn_pt = clean_name(row_dict.get('方案类型', ''))
            existing_txn = Transaction.query.filter_by(patient_id=patient.id, payment_date=txn_pd, plan_type=txn_pt, performance_amount=perf_amt).first()
            if not existing_txn:
                txn = Transaction(
                    patient_id=patient.id,
                    plan_type=txn_pt,
                    plan_detail=row_dict.get('方案（品牌、颗数、牙位）', ''),
                    brand=row_dict.get('品牌', ''),
                    tooth_positions=row_dict.get('牙位', ''),
                    tooth_count=parse_int(row_dict.get('颗数', '0')),
                    bone_graft=row_dict.get('骨粉', ''),
                    bone_membrane=row_dict.get('骨膜', ''),
                    sinus_lift=row_dict.get('内外提', ''),
                    deposit_amount=deposit_amt,
                    deposit_contract=contract_amt,
                    deposit_date=parse_date(row_dict.get('定金日期', '')),
                    deposit_return_date=parse_date(row_dict.get('定金回款日期', '')),
                    performance_amount=perf_amt,
                    paid_amount=paid_amt,
                    payment_date=txn_pd,
                    supplement_amount=supp_amt,
                    supplement_date=parse_date(row_dict.get('补费日期', '')),
                    debt_amount=debt_amt,
                    supplement_records=row_dict.get('补费记录（日期+金额）', ''),
                    treatment_date=parse_date(row_dict.get('治疗日期', '')),
                    treatment_doctor=clean_name(row_dict.get('治疗医生', '')),
                    consultant=clean_name(row_dict.get('谈单人', '')),
                    visit_outcome=clean_name(row_dict.get('成交类型', '')),
                )
                db.session.add(txn)
        
        imported += 1
    
    return imported


def import_from_files(flask_app=None):
    """Import all Excel data. Can be called with or without flask_app param."""
    ctx = flask_app.app_context() if flask_app else app.app_context()
    with ctx:
        db.create_all()
        
        if Patient.query.count() > 0:
            print('✅ Database already has data, skipping import')
            return
        
        print("=" * 50)
        print("📥 开始导入接诊业绩数据...")
        
        wb = openpyxl.load_workbook(os.path.join(DATA_DIR, '接诊业绩数据.xlsx'), data_only=True)
        total_perf = 0
        for sheet in wb.sheetnames:
            if sheet in wb.sheetnames:
                count = import_performance(wb, sheet)
                print(f"  {sheet}: {count} 条")
                total_perf += count
        
        print(f"\n📥 开始导入预约总表数据...")
        wb2 = openpyxl.load_workbook(os.path.join(DATA_DIR, '预约总表.xlsx'), data_only=True)
        total_appt = 0
        for sheet in wb2.sheetnames:
            if sheet in wb2.sheetnames:
                count = import_appointment(wb2, sheet)
                print(f"  {sheet}: {count} 条")
                total_appt += count
        
        db.session.commit()
        
        print("\n" + "=" * 50)
        print(f"✅ 导入完成!")
        print(f"   接诊业绩: {total_perf} 条记录")
        print(f"   预约总表: {total_appt} 条记录")
        
        patients = Patient.query.count()
        appointments = Appointment.query.count()
        transactions = Transaction.query.count()
        print(f"\n📊 数据库统计:")
        print(f"   患者: {patients} 人")
        print(f"   预约: {appointments} 条")
        print(f"   交易: {transactions} 条")


def main():
    from app import app as flask_app
    import_from_files(flask_app)

if __name__ == '__main__':
    main()
